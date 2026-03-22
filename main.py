import os
import pickle
import shutil
import yfinance as yf
import pandas as pd
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

PICKLE_DIR = "pickle_data"
OUTPUT_DIR = "reports"
TEMPLATE_PATH = "excel_templates/templ.xlsx"

ROWS = [
    "Total Revenue",
    "Cost Of Revenue",
    "Gross Profit",
    "Operating Expense",
    "Operating Income",
    "Net Income",
]

ROW_MAP = {
    "Total Revenue":     5,
    "Cost Of Revenue":   6,
    "Gross Profit":      7,
    "Operating Expense": 8,
    "Operating Income":  9,
    "Net Income":        10,
}


def fetch_financials(ticker_symbol: str) -> pd.DataFrame:
    path = os.path.join(PICKLE_DIR, f"{ticker_symbol}.pkl")
    if os.path.exists(path):
        with open(path, "rb") as f:
            return pickle.load(f)
    ticker = yf.Ticker(ticker_symbol)
    data = ticker.financials
    os.makedirs(PICKLE_DIR, exist_ok=True)
    with open(path, "wb") as f:
        pickle.dump(data, f)
    return data


def prepare_data(fin: pd.DataFrame) -> pd.DataFrame:
    available = [r for r in ROWS if r in fin.index]
    df = fin.loc[available]
    df.columns = [str(c.year) for c in df.columns]
    df = df / 1_000_000
    df = df.round(0).fillna(0).astype(int)
    return df


def generate_pdf(ticker_symbol: str, df: pd.DataFrame):
    out_path = os.path.join(OUTPUT_DIR, f"{ticker_symbol}_income_statement.pdf")
    if os.path.exists(out_path):
        print(f"PDF already exists, skipping: {out_path}")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    doc = SimpleDocTemplate(out_path, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(ticker_symbol, styles["Title"]))
    elements.append(Paragraph("Income Statement (USD millions)", styles["Normal"]))
    elements.append(Spacer(1, 20))

    years = list(df.columns)
    header = [""] + years
    data = [header]
    for row_name, row_data in df.iterrows():
        data.append([row_name] + [f"{v:,}" for v in row_data])

    table = Table(data, colWidths=[200] + [80] * len(years))
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table)
    doc.build(elements)
    print(f"Saved PDF: {out_path}")


def generate_excel(ticker_symbol: str, df: pd.DataFrame):
    out_path = os.path.join(OUTPUT_DIR, f"{ticker_symbol}_income_statement.xlsx")
    if os.path.exists(out_path):
        print(f"Excel already exists, skipping: {out_path}")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    shutil.copy2(TEMPLATE_PATH, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    ws["E1"] = ticker_symbol

    years = list(df.columns)
    data_cols = ["D", "E", "F", "G"]

    for i, year in enumerate(years[:4]):
        ws[f"{data_cols[i]}4"] = int(year)

    for row_name, excel_row in ROW_MAP.items():
        matched = [r for r in df.index if r.strip() == row_name.strip()]
        if not matched:
            continue
        for i, year in enumerate(years[:4]):
            ws[f"{data_cols[i]}{excel_row}"] = int(df.loc[matched[0], year])

    wb.save(out_path)
    print(f"Saved Excel: {out_path}")


def run(ticker_symbol: str, output: str = "pdf"):
    fin = fetch_financials(ticker_symbol)
    df = prepare_data(fin)
    if output == "pdf":
        generate_pdf(ticker_symbol, df)
    elif output == "excel":
        generate_excel(ticker_symbol, df)


run("AAPL", output="pdf")
run("AAPL", output="excel")
run("MSFT", output="pdf")
run("MSFT", output="excel")